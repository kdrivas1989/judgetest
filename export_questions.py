#!/usr/bin/env python3
"""Export all test questions to an Excel file with each test on a separate sheet."""

import sys
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from questions import TESTS

def export_to_excel(output_file="test_questions.xlsx"):
    """Export all tests to Excel with each test on a separate sheet."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Sort tests by chapter for logical ordering
    sorted_tests = sorted(TESTS.items(), key=lambda x: (x[1].get('chapter', ''), x[0]))

    for test_id, test_data in sorted_tests:
        # Create sheet name (max 31 chars for Excel)
        sheet_name = test_data['chapter']
        if '_regional' in test_id:
            sheet_name += ' Regional'
        elif '_national' in test_id:
            sheet_name += ' National'
        elif test_id == 'general':
            sheet_name = 'General (Ch 1-3)'

        # Truncate if too long
        sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # Add test info header
        ws.merge_cells('A1:G1')
        ws['A1'] = test_data['name']
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:G2')
        ws['A2'] = f"Passing Score: {test_data['passing_score']}%"
        ws['A2'].alignment = Alignment(horizontal="center")

        # Column headers
        headers = ['#', 'Question', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct', 'Section Reference']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add questions
        for row_idx, q in enumerate(test_data.get('questions', []), 5):
            # Question number
            ws.cell(row=row_idx, column=1, value=q['id']).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

            # Question text
            ws.cell(row=row_idx, column=2, value=q['question']).border = thin_border
            ws.cell(row=row_idx, column=2).alignment = cell_alignment

            # Options
            options = q.get('options', [])
            for opt_idx, option in enumerate(options):
                cell = ws.cell(row=row_idx, column=3 + opt_idx, value=option)
                cell.border = thin_border
                cell.alignment = cell_alignment
                # Highlight correct answer
                if opt_idx == q.get('correct'):
                    cell.fill = correct_fill

            # Correct answer letter
            correct_idx = q.get('correct', 0)
            correct_letter = chr(65 + correct_idx) if correct_idx < 4 else '?'
            ws.cell(row=row_idx, column=7, value=correct_letter).border = thin_border
            ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=7).fill = correct_fill

            # Section reference
            ws.cell(row=row_idx, column=8, value=q.get('correct_section', '')).border = thin_border
            ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal="center")

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15

    # Save workbook
    wb.save(output_file)
    print(f"Exported {len(TESTS)} tests to {output_file}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")

if __name__ == "__main__":
    export_to_excel()
